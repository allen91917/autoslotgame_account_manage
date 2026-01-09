from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os
import sys
import time

# ============================
# 建立 Chrome Driver（使用 ChromeDriverManager）
# ============================
def create_driver():
    """建立 Selenium ChromeDriver（使用 ChromeDriverManager 自動管理）"""

    print("使用 ChromeDriverManager 自動下載並管理 chromedriver...")
    driver_path = ChromeDriverManager().install()

    # ============================
    # Chrome Options
    # ============================
    chrome_options = Options()
    chrome_options.add_argument("--disable-infobars")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option("useAutomationExtension", False)
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_argument("--disable-save-password-bubble")

    # 關閉 Chrome 密碼儲存提示
    prefs = {
        "credentials_enable_service": False,
        "profile.password_manager_enabled": False
    }
    chrome_options.add_experimental_option("prefs", prefs)

    # 設定視窗大小
    chrome_options.add_argument("--window-size=1280,800")

    service = Service(driver_path)
    driver = webdriver.Chrome(service=service, options=chrome_options)

    # ============================
    # 最強 anti-detection
    # ============================
    driver.execute_cdp_cmd(
        "Page.addScriptToEvaluateOnNewDocument",
        {
            "source": """
                Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
            """
        },
    )

    return driver

# ============================
# 取得執行檔所在目錄（支援 PyInstaller 打包）
# ============================
def get_base_dir():
    """
    取得程式執行的基礎目錄
    如果是 PyInstaller 打包的 exe，會返回 exe 所在目錄
    如果是 Python 腳本，會返回腳本所在目錄
    """
    if getattr(sys, 'frozen', False):
        # 如果是打包後的 exe
        return os.path.dirname(sys.executable)
    else:
        # 如果是 Python 腳本
        return os.path.dirname(os.path.abspath(__file__))

# ============================
# 讀取用戶帳密 TXT
# ============================
def read_all_user_info():
    """
    讀取用戶資訊.txt 中的所有帳號密碼
    每一行格式： account,password
    忽略空行與 # 開頭的註解
    回傳 List[Tuple[str, str]]
    """
    base_dir = get_base_dir()
    txt_path = os.path.join(base_dir, "用戶資訊.txt")

    if not os.path.exists(txt_path):
        print(f"找不到 用戶資訊.txt")
        print(f"當前查找路徑: {txt_path}")
        print(f"exe 所在目錄: {base_dir}")
        raise FileNotFoundError("找不到 用戶資訊.txt")

    user_list = []

    with open(txt_path, "r", encoding="utf-8") as f:
        lines = f.readlines()

    for line in lines:
        line = line.strip()

        # ✅ 忽略空行與註解行
        if not line or line.startswith("#"):
            continue

        # ✅ 必須包含逗號
        if "," not in line:
            print(f"格式錯誤略過：{line}")
            continue

        account, password = line.split(",", 1)
        user_list.append((account.strip(), password.strip()))

    return user_list


def input_account_password(driver, account, password):
    """
    輸入指定帳密
    """
    wait = WebDriverWait(driver, 10)

    acc_input = wait.until(EC.presence_of_element_located(
        (By.XPATH, "//input[@id='form_item_account']")
    ))
    acc_input.clear()
    acc_input.send_keys(account)

    pwd_input = wait.until(EC.presence_of_element_located(
        (By.XPATH, "//input[@id='form_item_password']")
    ))
    pwd_input.clear()
    pwd_input.send_keys(password)
    time.sleep(1)  # 等待 1 秒，避免輸入過快被偵測 
    """
    自動點擊登入按鈕
    """
    wait = WebDriverWait(driver, 10)
    login_btn = wait.until(EC.element_to_be_clickable((
        By.XPATH,
        "//button/span[text()='登 錄']/parent::button"
    )))
    login_btn.click()


def close_announcement_popup(driver):
    """
    關閉公告彈窗
    """
    try:
        print("檢查是否有公告彈窗...")
        wait = WebDriverWait(driver, 10)
        close_button = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//button[.//span[text()='確 認']]"))
        )
        close_button.click()
        print("已關閉公告彈窗")
        time.sleep(1)
    except:
        print("未發現公告彈窗或已關閉")


def navigate_to_account_report(driver):
    """
    導航到帳務報表頁面
    """
    try:
        print("正在點擊「帳務報表」...")
        wait = WebDriverWait(driver, 20)
        
        # 點擊「帳務報表」選單
        account_report = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//span[contains(@class, 'vben-simple-menu-sub-title') and text()='帳務報表']"))
        )
        account_report.click()
        # print("已點擊「帳務報表」")
        time.sleep(2)
        
        # print("成功進入帳務報表頁面！")
        
    except Exception as e:
        print(f"❌ 導航到帳務報表失敗: {e}")
        raise

def extract_member_data(driver, login_account):
    """
    抓取遊戲會員報表資料
    回傳：包含所有會員資料的列表（排除登入帳號）
    """
    try:
        print("正在抓取報表資料...")
        time.sleep(3)  # 等待資料載入完成
        
        # 找到所有會員行
        member_rows = driver.find_elements(By.XPATH, "//div[contains(@class, 'my-table-row-box')]//div[contains(@class, 'my-table-row')]")
        
        data_list = []
        
        for index, row in enumerate(member_rows):
            try:
                # 抓取會員姓名和帳號
                name_account_element = row.find_element(By.XPATH, ".//div[contains(@class, 'my-table-cell')][1]//div")
                name_account_text = name_account_element.text.strip()
                
                # 如果是第一筆且包含登入帳號，則忽略
                if index == 0 and login_account in name_account_text:
                    # print(f"⚠️ 忽略第一筆（登入帳號）：{name_account_text}")
                    continue
                
                # 抓取注單筆數
                order_count = row.find_element(By.XPATH, ".//div[.//div[text()='注單筆數']]/div[2]").text.strip()
                
                # 抓取下注金額
                bet_amount = row.find_element(By.XPATH, ".//div[.//div[text()='下注金額']]/div[2]").text.strip()
                
                # 抓取有效投注
                valid_bet = row.find_element(By.XPATH, ".//div[.//div[text()='有效投注']]/div[2]").text.strip()
                
                # 抓取會員輸贏
                member_result = row.find_element(By.XPATH, ".//div[.//div[text()='會員輸贏']]/div[2]").text.strip()
                
                # 抓取會員退水
                member_rebate = row.find_element(By.XPATH, ".//div[.//div[text()='會員退水']]/div[2]").text.strip()
                
                # 抓取個人佔成
                personal_share = row.find_element(By.XPATH, ".//div[.//div[text()='個人佔成']]/div[2]").text.strip()
                
                # 抓取個人退水
                personal_rebate = row.find_element(By.XPATH, ".//div[.//div[text()='個人退水']]/div[2]").text.strip()
                
                # 抓取應收下線
                receivable = row.find_element(By.XPATH, ".//div[.//div[text()='應收下線']]/div[2]").text.strip()
                
                member_data = {
                    "姓名帳號": name_account_text,
                    "注單筆數": order_count,
                    "下注金額": bet_amount,
                    "有效投注": valid_bet,
                    "會員輸贏": member_result,
                    "會員退水": member_rebate,
                    "個人佔成": personal_share,
                    "個人退水": personal_rebate,
                    "應收下線": receivable
                }
                
                data_list.append(member_data)
                # print(f"✅ 已抓取：{name_account_text} - 應收下線: {receivable}")
                
            except Exception as e:
                print(f"抓取某筆資料時發生錯誤: {e}")
                continue
        
        return data_list
        
    except Exception as e:
        print(f"抓取報表資料失敗: {e}")
        import traceback
        traceback.print_exc()
        return []


def export_to_excel(report_data, login_account):
    """
    將報表資料匯出為 Excel 檔案到桌面（上週和本週放在同一個工作表）
    """
    try:
        # 獲取桌面路徑
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        file_name = f"{login_account}.xlsx"
        file_path = os.path.join(desktop_path, file_name)
        
        print(f"\n正在匯出 Excel 檔案到: {file_path}")
        
        # 創建新的工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "遊戲會員報表"
        
        # 標題樣式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 寫入標題（加上週期欄位）
        headers = ["週期", "姓名帳號", "注單筆數", "下注金額", "有效投注", "會員輸贏", "會員退水", "個人佔成", "個人退水", "應收下線"]
        ws.append(headers)
        
        # 設定標題樣式
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 寫入上週資料
        if report_data.get("上週"):
            last_week_total = 0
            for data in report_data["上週"]:
                ws.append([
                    "上週",
                    data["姓名帳號"],
                    data["注單筆數"],
                    data["下注金額"],
                    data["有效投注"],
                    data["會員輸贏"],
                    data["會員退水"],
                    data["個人佔成"],
                    data["個人退水"],
                    data["應收下線"]
                ])
                # 累加應收下線（處理逗號和正負號）
                try:
                    receivable_value = float(data["應收下線"].replace(",", ""))
                    last_week_total += receivable_value
                except:
                    pass
            
            # 加入上週總計行
            ws.append([
                "上週",
                "總應收下線",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                f"{last_week_total:,.2f}"
            ])
        
        # 空一行
        ws.append(["", "", "", "", "", "", "", "", "", ""])
        
        # 寫入本週資料
        if report_data.get("本週"):
            this_week_total = 0
            for data in report_data["本週"]:
                ws.append([
                    "本週",
                    data["姓名帳號"],
                    data["注單筆數"],
                    data["下注金額"],
                    data["有效投注"],
                    data["會員輸贏"],
                    data["會員退水"],
                    data["個人佔成"],
                    data["個人退水"],
                    data["應收下線"]
                ])
                # 累加應收下線（處理逗號和正負號）
                try:
                    receivable_value = float(data["應收下線"].replace(",", ""))
                    this_week_total += receivable_value
                except:
                    pass
            
            # 加入本週總計行
            ws.append([
                "本週",
                "總應收下線",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                f"{this_week_total:,.2f}"
            ])
        
        # 調整列寬
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # 儲存檔案
        wb.save(file_path)
        print(f"Excel 檔案已成功匯出: {file_path}")
        return file_path
        
    except Exception as e:
        print(f"❌ 匯出 Excel 失敗: {e}")
        import traceback
        traceback.print_exc()
        return None


def query_game_member_reports(driver, login_account):
    """
    查詢遊戲會員報表：點擊遊戲會員報表 → 上週查詢 → 本週查詢
    """
    # 定義絕對 XPath
    last_week_button_xpath = "/html/body/div[1]/section/section/section/div[3]/div/div/div/div/form/div/div[2]/div/div/div[2]/div/div/div/div/div/button[4]"
    this_week_button_xpath = "/html/body/div[1]/section/section/section/div[3]/div/div/div/div/form/div/div[2]/div/div/div[2]/div/div/div/div/div/button[3]"
    search_xpath = "/html/body/div[1]/section/section/section/div[3]/div/div/div/div/form/div/div[5]/div/div/div/div/div/div/button[2]"
    
    try:
        wait = WebDriverWait(driver, 30)
        
        # ===== 點擊「遊戲會員報表」 =====
        print("\n正在點擊「遊戲會員報表」...")
        game_member_report = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//label[contains(@class, 'ant-radio-button-wrapper')]//span[text()='遊戲會員報表']"))
        )
        game_member_report.click()
        # print("✅ 已點擊「遊戲會員報表」")
        time.sleep(3)
        
        # ===== 查詢上週 =====
        # print("\n正在點擊「上週」按鈕...")
        last_week_button = wait.until(
            EC.element_to_be_clickable((By.XPATH, last_week_button_xpath))
        )
        driver.execute_script("arguments[0].scrollIntoView(true);", last_week_button)
        time.sleep(1)
        driver.execute_script("arguments[0].click();", last_week_button)
        # print("✅ 已點擊「上週」")
        time.sleep(2)
        
        # print("正在點擊「查詢」按鈕(上週)...")
        query_button = wait.until(
            EC.element_to_be_clickable((By.XPATH, search_xpath))
        )
        driver.execute_script("arguments[0].scrollIntoView(true);", query_button)
        time.sleep(1)
        driver.execute_script("arguments[0].click();", query_button)
        # print("✅ 已執行上週查詢")
        time.sleep(5)
        
        # 抓取上週資料
        # print("\n開始抓取上週資料...")
        last_week_data = extract_member_data(driver, login_account)
        
        # ===== 查詢本週 =====
        # print("\n正在點擊「本週」按鈕...")
        this_week_button = wait.until(
            EC.element_to_be_clickable((By.XPATH, this_week_button_xpath))
        )
        driver.execute_script("arguments[0].scrollIntoView(true);", this_week_button)
        time.sleep(1)
        driver.execute_script("arguments[0].click();", this_week_button)
        # print("✅ 已點擊「本週」")
        time.sleep(2)
        
        # print("正在點擊「查詢」按鈕(本週)...")
        query_button = wait.until(
            EC.element_to_be_clickable((By.XPATH, search_xpath))
        )
        driver.execute_script("arguments[0].scrollIntoView(true);", query_button)
        time.sleep(1)
        driver.execute_script("arguments[0].click();", query_button)
        # print("✅ 已執行本週查詢")
        time.sleep(5)
        
        # 抓取本週資料
        # print("\n開始抓取本週資料...")
        this_week_data = extract_member_data(driver, login_account)
        
        # print("\n✅ 遊戲會員報表查詢完成！")
        
        return {
            "上週": last_week_data,
            "本週": this_week_data
        }
        
    except Exception as e:
        print(f"查詢遊戲會員報表失敗: {e}")
        import traceback
        traceback.print_exc()
        raise


if __name__ == "__main__":
    users = read_all_user_info()

    if not users:
        print("❌ 沒有任何有效帳號")
        sys.exit(1)

    # ✅ 從 txt 取得第一組帳號
    account, password = users[0]
    print(f"使用帳號登入：{account}")

    driver = create_driver()
    
    try:
        driver.get("https://admin.fin88.app")

        # ✅ 用 txt 的帳密登入
        input_account_password(driver, account, password)
        print("✅ 登入成功，等待頁面載入...")
        time.sleep(5)
        
        # ✅ 關閉公告彈窗
        close_announcement_popup(driver)
        
        # ✅ 導航到帳務報表
        navigate_to_account_report(driver)
        
        # ✅ 查詢遊戲會員報表（上週 + 本週）
        report_data = query_game_member_reports(driver, account)
        
        # ✅ 匯出為 Excel
        excel_path = export_to_excel(report_data, account)
        
        print("\n" + "="*50)
        print("所有操作完成！")
        print("="*50)
        print("5 秒後自動關閉瀏覽器並結束程式...")
        time.sleep(5)

        driver.quit()
        os._exit(0)
            
    except Exception as e:
        print(f"\n❌ 發生錯誤: {e}")
        

